from flask import Flask, render_template, request, redirect, url_for, send_file
from hdbcli import dbapi
from dotenv import load_dotenv
import os
import pandas as pd
import re
from io import BytesIO


load_dotenv()

app = Flask(__name__)

# SAP HANA Connection
conn = dbapi.connect(
    address=os.getenv("HANA_HOST"),
    port=int(os.getenv("HANA_PORT")),
    user=os.getenv("HANA_USER"),
    password=os.getenv("HANA_PASS")
)

SCHEMA = os.getenv("HANA_SCHEMA")
CAMPAIGN_TABLE = "ACH_FCA_CAMPAIGN"
LOOKUP_TABLE = "ACH_FCA_LOOKUP"
LOGS_TABLE = "ACH_FCA_LOGS"

# Column name mapping for user-friendly display
COLUMN_DISPLAY_NAMES = {
    'TENANTID': 'Tenant ID',
    'CAMPAIGNID': 'Campaign ID',
    'CREATEDATE': 'Create Date',
    'CAMPAIGNNAME': 'Campaign Name',
    'STARTDATE': 'Start Date',
    'ENDDATE': 'End Date',
    'STATUS': 'Status',
    'FCA': 'FCA',
    'IFCA': 'IFCA',
    'BVSHITS': 'BVS Hits',
    'SALESTYPE': 'Sales Type',
    'FCABUNDLERANGE': 'FCA Bundle Range',
    'RETSIMBUN': 'Retailer SIM Bundle',
    'BVSHITS_TO_FCA_RANGE': 'BVS Hits to FCA Range',
    'IFCADATERANGE': 'IFCA Date Range',
    'BUNDLEPRICETYPE': 'Bundle Price Type',
    'PRICETYPEVALUE': 'Price Type Value',
    'BUNDLE': 'Bundle',
    'RECHARGETYPE': 'Recharge Type',
    'BUNDLETYPE': 'Bundle Type',
    'RECHARGERNR': 'Recharge NR',
    'RECHARGERBR': 'Recharge BR',
    'RETAILERID': 'Retailer ID',
    'PRODUCTID': 'Product ID',
    'TARGET': 'Target',
    'COMMISSION': 'Commission',
    'MIN': 'Min',
    'MAX': 'Max',
    'CAP': 'Cap',
    'MODIFICATIONDATE': 'Modification Date'
}

# Function to convert column names to display names and handle special values
def get_display_name(column_name):
    return COLUMN_DISPLAY_NAMES.get(column_name, column_name)

# Function to convert 1/0 values to Yes/No for specific columns
def convert_yes_no(value, column_name):
    if column_name in ['FCA', 'IFCA', 'BVSHITS', 'BUNDLE']:
        if value == 1 or value == '1':
            return 'Yes'
        elif value == 0 or value == '0':
            return 'No'
        elif value is None:
            return ''
    return value

# -----------------------------
# Home / Welcome screen
# -----------------------------
@app.route("/")
def home():
    return render_template("home.html")

# -----------------------------
# Campaigns Table
# -----------------------------
@app.route("/campaigns")
def campaigns():
    cursor = conn.cursor()
    cursor.execute(f'SELECT * FROM "{SCHEMA}"."{CAMPAIGN_TABLE}" ORDER BY CAMPAIGNID')
    rows = cursor.fetchall()
    columns = [c[0] for c in cursor.description]

    # Create display columns - include ALL columns
    display_columns = [get_display_name(col) for col in columns]

    # Convert rows to lists for template and convert 1/0 to Yes/No
    extended_rows = []
    for row in rows:
        row_list = []
        for i, value in enumerate(row):
            converted_value = convert_yes_no(value, columns[i])
            row_list.append(converted_value)
        extended_rows.append(row_list)

    return render_template("campaigns.html", 
                         rows=extended_rows, 
                         columns=columns,
                         display_columns=display_columns, 
                         zip=zip)

@app.route("/campaigns/add", methods=["GET", "POST"])
def add_campaign():
    cursor = conn.cursor()

    if request.method == "POST":
        def empty_to_none(val):
            return None if val == '' else val
        
        def yes_no_to_int(val):
            # Convert Yes/No back to 1/0 for database
            if val == 'Yes' or val == '1':
                return 1
            elif val == 'No' or val == '0':
                return 0
            return empty_to_none(val)

        # Collect all form values except TENANTID, CAMPAIGNID, CREATEDATE
        data = {}
        for key, val in request.form.items():
            if key not in ['TENANTID', 'CAMPAIGNID', 'CREATEDATE']:
                # Convert Yes/No back to 1/0 for FCA, IFCA, BVSHITS, BUNDLE
                if key in ['FCA', 'IFCA', 'BVSHITS', 'BUNDLE']:
                    data[key] = yes_no_to_int(val)
                else:
                    data[key] = empty_to_none(val)

        # Clear recharge fields if RECHARGETYPE != RECHARGER
        if data.get('RECHARGETYPE') != 'RECHARGER':
            data['RECHARGERNR'] = None
            data['RECHARGERBR'] = None

        # Insert into database
        columns = list(data.keys())
        values = list(data.values())
        placeholders = ','.join(['?' for _ in columns])
        insert_stmt = f'INSERT INTO "{SCHEMA}"."{CAMPAIGN_TABLE}" ({",".join(columns)}) VALUES ({placeholders})'

        cursor.execute(insert_stmt, tuple(values))
        conn.commit()
        return redirect(url_for("campaigns"))

    # GET request: fetch columns excluding dependent fields
    cursor.execute(f'SELECT * FROM "{SCHEMA}"."{CAMPAIGN_TABLE}" LIMIT 1')
    all_columns = [c[0] for c in cursor.description]

    # Exclude dependent fields from main loop
    columns = [c for c in all_columns if c not in ['BUNDLE', 'RECHARGETYPE', 'BUNDLETYPE', 'RECHARGERNR', 'RECHARGERBR']]

    # Create display columns
    display_columns = [get_display_name(col) for col in columns]

    return render_template("add_campaign.html", 
                         columns=columns, 
                         display_columns=display_columns, 
                         zip=zip)

@app.route("/campaigns/edit/<int:campaignid>", methods=["GET", "POST"])
def edit_campaign(campaignid):
    cursor = conn.cursor()

    # Editable fields
    editable_fields = [
        "CAMPAIGNNAME",
        "STARTDATE",
        "ENDDATE",
        "STATUS",
        "FCABUNDLERANGE",
        "BVSHITS_TO_FCA_RANGE",
        "IFCADATERANGE",
        "RECHARGETYPE",
        "RECHARGERNR",
        "RECHARGERBR"
    ]

    if request.method == "POST":
        def empty_to_none(val):
            return None if val == '' else val

        # Collect only editable fields
        data = {}
        for field in editable_fields:
            data[field] = empty_to_none(request.form.get(field, None))

        # If RECHARGETYPE != RECHARGER, clear recharge fields
        if data.get('RECHARGETYPE') != 'RECHARGER':
            data['RECHARGERNR'] = None
            data['RECHARGERBR'] = None

        # Build update statement dynamically
        set_clause = ', '.join([f'"{col}"=?' for col in editable_fields])
        values = [data[col] for col in editable_fields]
        values.append(campaignid)  # for WHERE clause

        cursor.execute(f'''
            UPDATE "{SCHEMA}"."{CAMPAIGN_TABLE}"
            SET {set_clause}
            WHERE CAMPAIGNID=?
        ''', tuple(values))
        conn.commit()

        return redirect(url_for("campaigns"))

    # GET request: fetch existing campaign
    cursor.execute(f'SELECT * FROM "{SCHEMA}"."{CAMPAIGN_TABLE}" WHERE CAMPAIGNID=?', (campaignid,))
    campaign = cursor.fetchone()
    columns = [c[0] for c in cursor.description]
    
    # Convert row to dictionary for easier access in template and convert 1/0 to Yes/No
    campaign_dict = {}
    for i, col in enumerate(columns):
        campaign_dict[col] = convert_yes_no(campaign[i], col)

    # Create display columns
    display_columns = [get_display_name(col) for col in columns]

    return render_template("edit_campaign.html", 
                         campaign=campaign_dict, 
                         columns=columns, 
                         display_columns=display_columns,
                         zip=zip)

@app.route("/campaigns/delete/<int:campaignid>")
def delete_campaign(campaignid):
    cursor = conn.cursor()
    cursor.execute(f'DELETE FROM "{SCHEMA}"."{CAMPAIGN_TABLE}" WHERE CAMPAIGNID=?', (campaignid,))
    conn.commit()
    return redirect(url_for("campaigns"))

# -----------------------------
# Excel Upload/Download Routes
# -----------------------------
@app.route("/campaigns/download-template")
def download_campaign_template():
    """Download Excel template with instructions in right columns"""
    try:
        import io
        
        # Create main data DataFrame
        data = {
            'CAMPAIGNNAME': ['Sample Campaign', '', '', ''],
            'STARTDATE': ['2025-01-01', '', '', ''],
            'ENDDATE': ['2025-12-31', '', '', ''],
            'STATUS': ['1', '', '', ''],
            'FCA': ['1', '', '', ''],
            'IFCA': ['0', '', '', ''],
            'BVSHITS': ['1', '', '', ''],
            'BUNDLE': ['0', '', '', ''],
            'SALESTYPE': ['MNP', '', '', ''],
            'FCABUNDLERANGE': ['10', '', '', ''],
            'RETSIMBUN': ['', '', '', ''],
            'BVSHITS_TO_FCA_RANGE': ['5', '', '', ''],
            'IFCADATERANGE': ['10', '', '', ''],
            'BUNDLEPRICETYPE': ['RANGE', '', '', ''],
            'PRICETYPEVALUE': ['100-200;200-300;400-500', '', '', ''],
            'RECHARGETYPE': ['RECHARGER', '', '', ''],
            'BUNDLETYPE': ['POWER LOAD', '', '', ''],
            'RECHARGERNR': ['100', '', '', ''],
            'RECHARGERBR': ['100.5', '', '', '']
        }
        
        df = pd.DataFrame(data)
        
        output = io.BytesIO()
        
        # Create Excel file with openpyxl
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Template')
            
            # Get the worksheet
            workbook = writer.book
            worksheet = writer.sheets['Template']
            
            from openpyxl.styles import Font
            
            # STEP 1: Pre-format date columns as TEXT for ALL rows
            date_columns = ['STARTDATE', 'ENDDATE']
            date_col_indices = []
            
            # Find column indices for date columns
            for col_idx, col_name in enumerate(df.columns, 1):
                if col_name in date_columns:
                    date_col_indices.append(col_idx)
            
            # Apply text format to date columns for ALL rows
            for col_idx in date_col_indices:
                for row_idx in range(1, len(df) + 2):
                    worksheet.cell(row=row_idx, column=col_idx).number_format = '@'
            
            # STEP 2: Apply styling to main data area
            # Grey font for sample row (row 2)
            grey_font = Font(color="808080")
            for col in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=2, column=col)
                cell.font = grey_font
            
            # Bold headers (row 1)
            header_font = Font(bold=True)
            for col in range(1, len(df.columns) + 1):
                worksheet.cell(row=1, column=col).font = header_font
            
            # STEP 3: Add instructions in right columns
            # Add spacing (2 empty columns)
            spacer_col1 = len(df.columns) + 1
            spacer_col2 = len(df.columns) + 2
            
            # Instructions column
            instructions_col = len(df.columns) + 3
            
            # Add instruction headers
            worksheet.cell(row=1, column=instructions_col, value="INSTRUCTIONS").font = Font(bold=True, color="FF0000")
            
            # Add instructions for each row
            instructions_data = {
                2: "SAMPLE DATA - This row will be ignored during upload",
                3: "ENTER YOUR DATA HERE Required fields: Campaign Name, Start Date, End Date, STATUS",
                4: "FCA, IFCA, BVSHITS, BUNDLE: Use 1 for Yes, 0 for No",
                5: "DATES: Use YYYY-MM-DD format ",
                6: "STATUS: Use 1 for Active, 0 for Inactive",
                7: "SALESTYPE: MNP, BYN, NPP, MNPBVS_NEW, e_SIM_BYN, D2C",
                8: "RECHARGETYPE: RECHARGER, BUNDLE, NORMAL RECHARGE, ALL",
                9: "BUNDLETYPE: POWER LOAD, DIGITAL, ADC, FS, ALL",
                10: "Clear RECHARGERNR & RECHARGERBR if RECHARGETYPE IS NOT EQULAS TO RECHARGER",
                11: "PRICETYPEVALUE: Use semicolons for multiple ranges (100-200;200-300)",
                12: "Dates must be valid (e.g., no 2025-06-31 - June has 30 days)",
                13:"Add more campaigns below",
                14:"Save file before uploading",
            }


            
            for row, instruction in instructions_data.items():
                worksheet.cell(row=row, column=instructions_col, value=instruction)
            
            # Set column widths for better visibility
            worksheet.column_dimensions[chr(64 + instructions_col)].width = 50
        
        output.seek(0)
        
        return send_file(
            output,
            as_attachment=True,
            download_name='campaign_template.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        return f"Error generating template: {str(e)}", 500
    
    
@app.route("/campaigns/upload", methods=["GET", "POST"])
def upload_campaigns():
    """Handle Excel file upload - ignore instruction columns safely"""
    if request.method == "GET":
        return render_template("upload_campaign.html")

    if request.method == "POST":
        try:
            if 'file' not in request.files:
                return "No file uploaded", 400

            file = request.files['file']
            if file.filename == '':
                return "No file selected", 400

            if not file.filename.endswith(('.xlsx', '.xls')):
                return "Please upload an Excel file", 400

            # Expected REAL columns only (19)
            expected_columns = [
                'CAMPAIGNNAME', 'STARTDATE', 'ENDDATE', 'STATUS', 'FCA', 'IFCA',
                'BVSHITS', 'BUNDLE', 'SALESTYPE', 'FCABUNDLERANGE', 'RETSIMBUN',
                'BVSHITS_TO_FCA_RANGE', 'IFCADATERANGE', 'BUNDLEPRICETYPE',
                'PRICETYPEVALUE', 'RECHARGETYPE', 'BUNDLETYPE',
                'RECHARGERNR', 'RECHARGERBR'
            ]

            # Read full Excel sheet but as strings
            df = pd.read_excel(file, dtype=str)

            print("=== DEBUG: Original Columns ===")
            print(df.columns.tolist())
            print("===============================")

            # Force pandas to keep ONLY the valid 19 columns
            # (Ignore any instruction / extra columns safely)
            df = df.reindex(columns=expected_columns)

            print("=== DEBUG: After Filtering to Expected Columns ===")
            print(df.columns.tolist())
            print("===============================")

            # Skip sample row & empty rows
            df = df.iloc[1:]
            df = df.dropna(how='all')

            success_count = 0
            error_count = 0
            errors = []
            cursor = conn.cursor()

            for index, row in df.iterrows():
                if row.isna().all():
                    continue

                data = {}

                # Loop through expected columns
                for col in expected_columns:
                    value = row[col]

                    # Normalize nulls
                    if pd.isna(value) or str(value).strip() in ['', 'nan', 'None']:
                        value = None
                    else:
                        value = str(value).strip()

                    # Convert status
                    if col == 'STATUS' and value is not None:
                        v = value.lower()
                        if v in ['1', 'active', 'yes', 'true', 'y']:
                            value = 1
                        elif v in ['0', 'inactive', 'no', 'false', 'n']:
                            value = 0
                        else:
                            value = None

                    # Convert FCA/IFCA/BVSHITS/BUNDLE
                    if col in ['FCA', 'IFCA', 'BVSHITS', 'BUNDLE'] and value is not None:
                        v = value.lower()
                        if v in ['1', 'yes', 'true', 'y']:
                            value = 1
                        elif v in ['0', 'no', 'false', 'n']:
                            value = 0
                        else:
                            value = None

                    # Convert date formats (YYYY-MM-DD)
                    if col in ['STARTDATE', 'ENDDATE'] and value is not None:
                        try:
                            parsed = pd.to_datetime(value)
                            value = parsed.strftime('%Y-%m-%d')
                        except:
                            pass

                    data[col] = value

                # Required field validations
                if not data['CAMPAIGNNAME']:
                    errors.append(f"Row {index+3}: CAMPAIGNNAME is required")
                    error_count += 1
                    continue

                if not data['STARTDATE']:
                    errors.append(f"Row {index+3}: STARTDATE is required")
                    error_count += 1
                    continue

                if not data['ENDDATE']:
                    errors.append(f"Row {index+3}: ENDDATE is required")
                    error_count += 1
                    continue

                if data['STATUS'] is None:
                    errors.append(f"Row {index+3}: STATUS is required (use 1 or 0)")
                    error_count += 1
                    continue

                # Clear RECHARGER fields if RECHARGETYPE != RECHARGER
                if data.get('RECHARGETYPE') != 'RECHARGER':
                    data['RECHARGERNR'] = None
                    data['RECHARGERBR'] = None

                # Insert into database
                cols = list(data.keys())
                vals = list(data.values())
                placeholders = ",".join(["?" for _ in cols])

                query = f'INSERT INTO "{SCHEMA}"."{CAMPAIGN_TABLE}" ({",".join(cols)}) VALUES ({placeholders})'
                cursor.execute(query, tuple(vals))
                success_count += 1

            conn.commit()

            # Response message
            result = f"Imported: {success_count}, Errors: {error_count}"
            if errors:
                result += "<br><br>" + "<br>".join(errors)

            return render_template("upload_campaign.html",
                                   result_message=result,
                                   success_count=success_count,
                                   error_count=error_count)

        except Exception as e:
            conn.rollback()
            return render_template("upload_campaign.html",
                                   result_message=f"Upload failed: {str(e)}")

# -----------------------------
# Lookup Table
# -----------------------------
@app.route("/lookup")
def lookup():
    cursor = conn.cursor()
    cursor.execute(f'SELECT * FROM "{SCHEMA}"."{LOOKUP_TABLE}" ORDER BY CAMPAIGNID')
    rows = cursor.fetchall()
    columns = [c[0] for c in cursor.description]
    
    # Create display columns
    display_columns = [get_display_name(col) for col in columns]
    
    return render_template("lookup.html", 
                         rows=rows, 
                         columns=columns, 
                         display_columns=display_columns,
                         zip=zip)

@app.route("/lookup/add", methods=["GET", "POST"])
def add_lookup():
    if request.method == "POST":
        campaignid = request.form["CAMPAIGNID"]
        retailerid = request.form["RETAILERID"]
        productid = request.form["PRODUCTID"]  # Now required
        startdate = request.form.get("STARTDATE") or None
        enddate = request.form.get("ENDDATE") or None
        
        # Convert empty strings to None for numeric fields
        target = request.form.get("TARGET")
        target = int(target) if target and target.strip() else None
        
        commission = request.form.get("COMMISSION") 
        commission = float(commission) if commission and commission.strip() else None
        
        min_val = request.form.get("MIN")
        min_val = float(min_val) if min_val and min_val.strip() else None
        
        max_val = request.form.get("MAX")
        max_val = float(max_val) if max_val and max_val.strip() else None
        
        cap = request.form.get("CAP")
        cap = float(cap) if cap and cap.strip() else None

        # Validate required fields
        if not campaignid or not retailerid or not productid:
            return "CAMPAIGNID, RETAILERID and PRODUCTID are required fields", 400

        cursor = conn.cursor()
        cursor.execute(f'''
            INSERT INTO "{SCHEMA}"."{LOOKUP_TABLE}"
            (CAMPAIGNID, RETAILERID, PRODUCTID, STARTDATE, ENDDATE, TARGET, COMMISSION, MIN, MAX, CAP)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (campaignid, retailerid, productid, startdate, enddate, target, commission, min_val, max_val, cap))
        conn.commit()
        return redirect(url_for("lookup"))
    
    # For GET request, show the form with display names
    cursor = conn.cursor()
    cursor.execute(f'SELECT * FROM "{SCHEMA}"."{LOOKUP_TABLE}" LIMIT 1')
    columns = [c[0] for c in cursor.description]
    display_columns = [get_display_name(col) for col in columns]
    
    return render_template("add_lookup.html", 
                         columns=columns, 
                         display_columns=display_columns,
                         zip=zip)


# Lookup Excel Upload/Download Routes
@app.route("/lookup/download-template")
def download_lookup_template():
    """Download Excel template with instructions in right columns"""
    try:
        import io
        
        # Create main data DataFrame
        data = {
            'CAMPAIGNID': ['1', '', '', ''],
            'RETAILERID': ['RET001', '', '', ''],
            'PRODUCTID': ['PROD001', '', '', ''],
            'STARTDATE': ['2025-01-01', '', '', ''],
            'ENDDATE': ['2025-12-31', '', '', ''],
            'TARGET': ['100', '', '', ''],
            'COMMISSION': ['50', '', '', ''],
            'MIN': ['1', '', '', ''],
            'MAX': ['10', '', '', ''],
            'CAP': ['1000', '', '', '']
        }
        
        df = pd.DataFrame(data)
        
        output = io.BytesIO()
        
        # Create Excel file with openpyxl
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Template')
            
            # Get the worksheet
            workbook = writer.book
            worksheet = writer.sheets['Template']
            
            from openpyxl.styles import Font
            
            # STEP 1: Pre-format date columns as TEXT for ALL rows
            date_columns = ['STARTDATE', 'ENDDATE']
            date_col_indices = []
            
            # Find column indices for date columns
            for col_idx, col_name in enumerate(df.columns, 1):
                if col_name in date_columns:
                    date_col_indices.append(col_idx)
            
            # Apply text format to date columns for ALL rows
            for col_idx in date_col_indices:
                for row_idx in range(1, len(df) + 2):
                    worksheet.cell(row=row_idx, column=col_idx).number_format = '@'
            
            # STEP 2: Apply styling to main data area
            # Grey font for sample row (row 2)
            grey_font = Font(color="808080")
            for col in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=2, column=col)
                cell.font = grey_font
            
            # Bold headers (row 1)
            header_font = Font(bold=True)
            for col in range(1, len(df.columns) + 1):
                worksheet.cell(row=1, column=col).font = header_font
            
            # STEP 3: Add instructions in right columns
            # Add spacing (2 empty columns)
            spacer_col1 = len(df.columns) + 1
            spacer_col2 = len(df.columns) + 2
            
            # Instructions column
            instructions_col = len(df.columns) + 3
            
            # Add instruction headers
            worksheet.cell(row=1, column=instructions_col, value="INSTRUCTIONS").font = Font(bold=True, color="FF0000")
            
            # Add instructions for each row
            instructions_data = {
                2: "SAMPLE DATA - This row will be ignored during upload",
                3: "ENTER YOUR DATA HERE - Required fields: CAMPAIGNID, RETAILERID, PRODUCTID",
                4: "CAMPAIGNID: Must exist in campaigns table",
                5: "RETAILERID: Unique retailer identifier",
                6: "PRODUCTID: Unique product identifier",
                7: "DATES: Use YYYY-MM-DD format",
                8: "TARGET, COMMISSION, MIN, MAX, CAP: Use numeric values",
                9: "CAMPAIGNID + RETAILERID + PRODUCTID must be unique",
                10: "Dates must be valid (e.g., no 2025-06-31 - June has 30 days)",
                11: "Add more lookup entries below",
                12: "Save file before uploading",
            }
            
            for row, instruction in instructions_data.items():
                worksheet.cell(row=row, column=instructions_col, value=instruction)
            
            # Set column widths for better visibility
            worksheet.column_dimensions[chr(64 + instructions_col)].width = 50
        
        output.seek(0)
        
        return send_file(
            output,
            as_attachment=True,
            download_name='lookup_template.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        return f"Error generating template: {str(e)}", 500
    
    
@app.route("/lookup/upload", methods=["GET", "POST"])
def upload_lookup():
    """Handle Excel file upload - ignore instruction columns safely"""
    if request.method == "GET":
        return render_template("upload_lookup.html")

    if request.method == "POST":
        try:
            if 'file' not in request.files:
                return "No file uploaded", 400

            file = request.files['file']
            if file.filename == '':
                return "No file selected", 400

            if not file.filename.endswith(('.xlsx', '.xls')):
                return "Please upload an Excel file", 400

            # Expected REAL columns only (10)
            expected_columns = [
                'CAMPAIGNID', 'RETAILERID', 'PRODUCTID', 'STARTDATE', 'ENDDATE',
                'TARGET', 'COMMISSION', 'MIN', 'MAX', 'CAP'
            ]

            # Read full Excel sheet but as strings
            df = pd.read_excel(file, dtype=str)

            print("=== DEBUG: Original Columns ===")
            print(df.columns.tolist())
            print("===============================")

            # Force pandas to keep ONLY the valid 10 columns
            # (Ignore any instruction / extra columns safely)
            df = df.reindex(columns=expected_columns)

            print("=== DEBUG: After Filtering to Expected Columns ===")
            print(df.columns.tolist())
            print("===============================")

            # Skip sample row & empty rows
            df = df.iloc[1:]
            df = df.dropna(how='all')

            success_count = 0
            error_count = 0
            errors = []
            cursor = conn.cursor()

            for index, row in df.iterrows():
                if row.isna().all():
                    continue

                data = {}

                # Loop through expected columns
                for col in expected_columns:
                    value = row[col]

                    # Normalize nulls
                    if pd.isna(value) or str(value).strip() in ['', 'nan', 'None']:
                        value = None
                    else:
                        value = str(value).strip()

                    # Convert numeric fields
                    if col in ['CAMPAIGNID', 'TARGET', 'COMMISSION', 'MIN', 'MAX', 'CAP'] and value is not None:
                        try:
                            value = int(float(value))
                        except (ValueError, TypeError):
                            value = None

                    # Convert date formats (YYYY-MM-DD)
                    if col in ['STARTDATE', 'ENDDATE'] and value is not None:
                        try:
                            parsed = pd.to_datetime(value)
                            value = parsed.strftime('%Y-%m-%d')
                        except:
                            pass

                    data[col] = value

                # Required field validations - ONLY these 3 are required
                if not data['CAMPAIGNID']:
                    errors.append(f"Row {index+3}: CAMPAIGNID is required")
                    error_count += 1
                    continue

                if not data['RETAILERID']:
                    errors.append(f"Row {index+3}: RETAILERID is required")
                    error_count += 1
                    continue

                if not data['PRODUCTID']:
                    errors.append(f"Row {index+3}: PRODUCTID is required")
                    error_count += 1
                    continue

                # All other fields (STARTDATE, ENDDATE, TARGET, etc.) are optional
                # They can be None/empty

                # Insert into database
                cols = list(data.keys())
                vals = list(data.values())
                placeholders = ",".join(["?" for _ in cols])

                query = f'INSERT INTO "{SCHEMA}"."{LOOKUP_TABLE}" ({",".join(cols)}) VALUES ({placeholders})'
                cursor.execute(query, tuple(vals))
                success_count += 1

            conn.commit()

            # Response message
            result = f"Imported: {success_count}, Errors: {error_count}"
            if errors:
                result += "<br><br>" + "<br>".join(errors)

            return render_template("upload_lookup.html",
                                   result_message=result,
                                   success_count=success_count,
                                   error_count=error_count)

        except Exception as e:
            conn.rollback()
            return render_template("upload_lookup.html",
                                   result_message=f"Upload failed: {str(e)}")



@app.route("/lookup/edit/<int:campaignid>/<retailerid>/<productid>", methods=["GET", "POST"])
def edit_lookup(campaignid, retailerid, productid):
    cursor = conn.cursor()

    # Helper function to convert empty strings to None for numeric columns
    def empty_to_none(value):
        return None if value == '' else value

    if request.method == "POST":
        # Get form values
        tenantid = request.form.get("TENANTID")  # read-only, but can still include if needed
        new_campaignid = request.form.get("CAMPAIGNID")  # editable
        startdate = request.form.get("STARTDATE")
        enddate = request.form.get("ENDDATE")
        target = empty_to_none(request.form.get("TARGET"))
        commission = empty_to_none(request.form.get("COMMISSION"))
        min_val = empty_to_none(request.form.get("MIN"))
        max_val = empty_to_none(request.form.get("MAX"))
        cap = empty_to_none(request.form.get("CAP"))

        # Update the lookup table; update CAMPAIGNID as well
        cursor.execute(f'''
            UPDATE "{SCHEMA}"."{LOOKUP_TABLE}"
            SET CAMPAIGNID=?, STARTDATE=?, ENDDATE=?, TARGET=?, COMMISSION=?, MIN=?, MAX=?, CAP=?, MODIFICATIONDATE=CURRENT_TIMESTAMP
            WHERE CAMPAIGNID=? AND RETAILERID=? AND PRODUCTID=?
        ''', (new_campaignid, startdate, enddate, target, commission, min_val, max_val, cap, campaignid, retailerid, productid))
        conn.commit()

        return redirect(url_for("lookup"))

    # GET request: fetch the existing row
    cursor.execute(f'''
        SELECT * FROM "{SCHEMA}"."{LOOKUP_TABLE}"
        WHERE CAMPAIGNID=? AND RETAILERID=? AND PRODUCTID=?
    ''', (campaignid, retailerid, productid))
    lookup_row = cursor.fetchone()
    columns = [c[0] for c in cursor.description]

    # Create display columns
    display_columns = [get_display_name(col) for col in columns]

    return render_template("edit_lookup.html", 
                         lookup=lookup_row, 
                         columns=columns, 
                         display_columns=display_columns,
                         zip=zip)

@app.route("/lookup/delete/<int:campaignid>/<retailerid>/<productid>")
def delete_lookup(campaignid, retailerid, productid):
    cursor = conn.cursor()
    cursor.execute(f'''
        DELETE FROM "{SCHEMA}"."{LOOKUP_TABLE}" 
        WHERE CAMPAIGNID=? AND RETAILERID=? AND PRODUCTID=?
    ''', (campaignid, retailerid, productid))
    conn.commit()
    return redirect(url_for("lookup"))

#BULK DELETE

@app.route('/lookup/delete_bulk/<int:campaign_id>')
def delete_bulk_lookup(campaign_id):
    cursor = conn.cursor()
    cursor.execute(f'''
        DELETE FROM "{SCHEMA}"."{LOOKUP_TABLE}" WHERE CAMPAIGNID=?
    ''', (campaign_id,))
    conn.commit()
    return redirect(url_for('lookup'))



# -----------------------------
# Logs Table
# -----------------------------
@app.route("/logs")
def logs():
    cursor = conn.cursor()
    cursor.execute(f'SELECT * FROM "{SCHEMA}"."{LOGS_TABLE}" ORDER BY COMPENSATIONDATE DESC')
    rows = cursor.fetchall()
    columns = [c[0] for c in cursor.description]
    
    # Create display columns
    display_columns = [get_display_name(col) for col in columns]
    
    return render_template("logs.html", 
                         rows=rows, 
                         columns=columns, 
                         display_columns=display_columns,
                         zip=zip)

# -----------------------------
# Run server
# -----------------------------
if __name__ == "__main__":
    app.run(debug=True)